"""
reporting/excel_reporter.py
----------------------------
Generates a professional, multi-sheet Excel report.

Sheets:
  1. Summary Dashboard
  2. Equity Curves
  3. Trade Log
  4. Risk Metrics
  5. Monthly Returns
  6. Drawdown Analysis
  7. Paper Trading — Session Overview
  8. Paper Trading — Order Execution Log
  9. Paper Trading — Live Positions
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import datetime


# ── Colour palette ─────────────────────────────────────────────────────────────
C = {
    "navy":        "1B2A4A",
    "blue":        "2563EB",
    "light_blue":  "DBEAFE",
    "green":       "16A34A",
    "light_green": "DCFCE7",
    "red":         "DC2626",
    "light_red":   "FEE2E2",
    "gold":        "D97706",
    "light_gold":  "FEF3C7",
    "grey":        "6B7280",
    "light_grey":  "F3F4F6",
    "white":       "FFFFFF",
    "black":       "111827",
    "input_blue":  "0000FF",
}

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _font(bold=False, color="111827", size=10, italic=False) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def _header_row(ws, row: int, headers: list, fill_color: str = "1B2A4A",
                font_color: str = "FFFFFF", start_col: int = 1):
    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.font      = _font(bold=True, color=font_color, size=10)
        c.fill      = _fill(fill_color)
        c.alignment = _align("center")
        c.border    = BORDER


def _section_title(ws, row: int, col: int, title: str):
    c = ws.cell(row=row, column=col, value=title)
    c.font      = _font(bold=True, color="1B2A4A", size=12)
    c.alignment = _align()


# ══════════════════════════════════════════════════════════════════════════════
# Main entry point
# ══════════════════════════════════════════════════════════════════════════════

def generate_report(
    results: dict,                  # {strategy_name: backtest_result_dict}
    metrics: dict,                  # {strategy_name: risk_metrics_dict}
    paper_results: dict = None,     # {strategy_name: paper_trading_result_dict}
    output_path: str = "quant_report.xlsx",
):
    wb = Workbook()
    wb.remove(wb.active)    # remove default blank sheet

    _make_summary(wb, results, metrics)
    _make_equity_curves(wb, results)
    _make_trade_log(wb, results)
    _make_risk_metrics(wb, metrics)
    _make_monthly_returns(wb, results)
    _make_drawdown(wb, results)

    if paper_results:
        _make_paper_overview(wb, paper_results)
        _make_paper_executions(wb, paper_results)
        _make_paper_positions(wb, paper_results)

    wb.save(output_path)
    print(f"[Reporter] Report saved → {output_path}")
    return output_path


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: Summary Dashboard
# ══════════════════════════════════════════════════════════════════════════════

def _make_summary(wb: Workbook, results: dict, metrics: dict):
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "📈  Quantitative Trading System — Performance Dashboard"
    c.font      = _font(bold=True, color="FFFFFF", size=14)
    c.fill      = _fill("1B2A4A")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:J2")
    ts = ws["A2"]
    ts.value     = f"Generated: {datetime.datetime.now().strftime('%d %b %Y %H:%M')}   |   Backtest Universe: AAPL, MSFT, GOOGL, AMZN, TSLA, SPY, QQQ, NVDA"
    ts.font      = _font(color="6B7280", size=9, italic=True)
    ts.alignment = _align("center")
    ts.fill      = _fill("F9FAFB")

    # ── Strategy Comparison Table ──────────────────────────────────────────────
    _section_title(ws, 4, 1, "Strategy Performance Comparison")

    headers = [
        "Strategy", "Total Return (%)", "Ann. Return (%)", "Ann. Vol (%)",
        "Sharpe Ratio", "Sortino Ratio", "Max Drawdown (%)",
        "VaR 95% (%)", "Win Rate (%)", "Total Trades",
    ]
    _header_row(ws, 5, headers)

    best_sharpe = max(metrics, key=lambda k: metrics[k].get("Sharpe Ratio", -999))

    for row_i, (strat, m) in enumerate(metrics.items(), start=6):
        vals = [
            strat,
            round((results[strat]["total_return"]) * 100, 2),
            m.get("Annualised Return (%)", ""),
            m.get("Annualised Volatility (%)", ""),
            m.get("Sharpe Ratio", ""),
            m.get("Sortino Ratio", ""),
            m.get("Max Drawdown (%)", ""),
            m.get("VaR 95% (daily %)", ""),
            m.get("Win Rate (%)", ""),
            m.get("Total Trades", ""),
        ]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border    = BORDER
            c.alignment = _align("center")
            bg = "DBEAFE" if strat == best_sharpe else ("FFFFFF" if row_i % 2 == 0 else "F9FAFB")
            c.fill = _fill(bg)
            if col_i == 1:
                c.font = _font(bold=(strat == best_sharpe), color="1B2A4A")
            else:
                c.font = _font(color="111827")

    # Best strategy callout
    best_row = 6 + len(metrics) + 1
    ws.merge_cells(f"A{best_row}:J{best_row}")
    star = ws.cell(row=best_row, column=1,
                   value=f"⭐  Best Risk-Adjusted Strategy: {best_sharpe}  (Sharpe: {metrics[best_sharpe]['Sharpe Ratio']})")
    star.font      = _font(bold=True, color="D97706", size=11)
    star.fill      = _fill("FEF3C7")
    star.alignment = _align("center")

    # Column widths
    widths = [24, 18, 18, 18, 14, 14, 18, 16, 14, 14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: Equity Curves
# ══════════════════════════════════════════════════════════════════════════════

def _make_equity_curves(wb: Workbook, results: dict):
    ws = wb.create_sheet("Equity Curves")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "Portfolio Equity Curves  — Starting Capital: $1,000,000"
    c.font      = _font(bold=True, color="FFFFFF", size=13)
    c.fill      = _fill("1B2A4A")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # Build combined DataFrame
    curves = {}
    for strat, res in results.items():
        curves[strat] = res["equity_curve"]
    df = pd.DataFrame(curves)
    df.index = pd.DatetimeIndex(df.index).normalize()

    headers = ["Date"] + list(df.columns)
    _header_row(ws, 3, headers)

    strat_colors = ["2563EB", "16A34A", "DC2626", "D97706", "7C3AED"]

    for row_i, (date, row) in enumerate(df.iterrows(), start=4):
        ws.cell(row=row_i, column=1, value=date.date()).border = BORDER
        ws.cell(row=row_i, column=1).font = _font(size=9)
        ws.cell(row=row_i, column=1).fill = _fill("F9FAFB")
        for col_i, val in enumerate(row, start=2):
            c = ws.cell(row=row_i, column=col_i, value=round(float(val), 2))
            c.number_format = '$#,##0'
            c.border  = BORDER
            c.font    = _font(size=9, color=strat_colors[(col_i - 2) % len(strat_colors)])
            c.fill    = _fill("FFFFFF" if row_i % 2 == 0 else "F9FAFB")
            c.alignment = _align("right")

    ws.column_dimensions["A"].width = 14
    for i in range(2, len(headers) + 1):
        _set_col_width(ws, i, 20)

    # Freeze header
    ws.freeze_panes = "A4"


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: Trade Log
# ══════════════════════════════════════════════════════════════════════════════

def _make_trade_log(wb: Workbook, results: dict):
    ws = wb.create_sheet("Trade Log")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "Complete Trade Log — All Strategies"
    c.font      = _font(bold=True, color="FFFFFF", size=13)
    c.fill      = _fill("1B2A4A")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    headers = ["Strategy", "Ticker", "Entry Date", "Exit Date",
               "Entry Price", "Exit Price", "Return (%)", "P&L ($)",
               "Duration (days)", "Result"]
    _header_row(ws, 3, headers)

    row_i = 4
    for strat, res in results.items():
        tlog = res["trade_log"].copy()
        if tlog.empty:
            continue
        tlog.insert(0, "Strategy", strat)
        for _, trade in tlog.iterrows():
            is_win = trade["Result"] == "Win"
            bg = "DCFCE7" if (row_i % 2 == 0 and is_win) else \
                 "FEE2E2" if (row_i % 2 == 0 and not is_win) else \
                 "F0FDF4" if is_win else "FFF1F2"
            for col_i, col in enumerate(headers, start=1):
                val = trade[col] if col != "Strategy" else strat
                c = ws.cell(row=row_i, column=col_i, value=val)
                c.fill      = _fill(bg)
                c.border    = BORDER
                c.alignment = _align("center")
                c.font      = _font(size=9)
                if col == "Return (%)":
                    c.number_format = "0.00%"
                    c.value = float(trade["Return (%)"]) / 100
                elif col == "P&L ($)":
                    c.number_format = '$#,##0.00'
                elif col in ("Entry Price", "Exit Price"):
                    c.number_format = '$#,##0.00'
            row_i += 1

    ws.freeze_panes = "A4"
    widths = [22, 8, 12, 12, 13, 13, 12, 14, 16, 8]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4: Risk Metrics
# ══════════════════════════════════════════════════════════════════════════════

def _make_risk_metrics(wb: Workbook, metrics: dict):
    ws = wb.create_sheet("Risk Metrics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = "Risk & Performance Metrics — Strategy Breakdown"
    c.font      = _font(bold=True, color="FFFFFF", size=13)
    c.fill      = _fill("1B2A4A")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    row_i = 3
    for strat, m in metrics.items():
        # Strategy sub-header
        ws.merge_cells(f"A{row_i}:C{row_i}")
        sh = ws.cell(row=row_i, column=1, value=f"◆  {strat}")
        sh.font      = _font(bold=True, color="FFFFFF", size=11)
        sh.fill      = _fill("2563EB")
        sh.alignment = _align()
        row_i += 1

        _header_row(ws, row_i, ["Metric", "Value", "Interpretation"],
                    fill_color="DBEAFE", font_color="1B2A4A")
        row_i += 1

        interpretations = {
            "Annualised Return (%)":  "Higher is better; benchmark S&P 500 ≈ 10%/yr",
            "Annualised Volatility (%)": "Lower is better; S&P 500 ≈ 15%/yr",
            "Sharpe Ratio":           "> 1.0 good; > 2.0 excellent",
            "Sortino Ratio":          "> 1.5 good; penalises downside only",
            "Max Drawdown (%)":       "Closer to 0 is better; max peak-to-trough loss",
            "VaR 95% (daily %)":      "Daily loss not exceeded 95% of trading days",
            "CVaR 95% (daily %)":     "Average loss on worst 5% of days (tail risk)",
            "Calmar Ratio":           "Ann. return / Max drawdown; > 0.5 acceptable",
            "Win Rate (%)":           "% of trades that are profitable",
            "Total Trades":           "Total closed trades over backtest period",
            "Profit Factor":          "Gross profit / Gross loss; > 1.5 is good",
            "Max DD Start":           "Start of worst drawdown period",
            "Max DD End":             "End (recovery) of worst drawdown period",
        }

        for metric, val in m.items():
            bg = "F9FAFB" if row_i % 2 == 0 else "FFFFFF"
            ws.cell(row=row_i, column=1, value=metric).fill  = _fill(bg)
            ws.cell(row=row_i, column=1).font      = _font(bold=True, size=9)
            ws.cell(row=row_i, column=1).border    = BORDER
            ws.cell(row=row_i, column=2, value=val).fill     = _fill(bg)
            ws.cell(row=row_i, column=2).font      = _font(size=9, color="2563EB")
            ws.cell(row=row_i, column=2).border    = BORDER
            ws.cell(row=row_i, column=2).alignment = _align("center")
            interp = interpretations.get(metric, "")
            ws.cell(row=row_i, column=3, value=interp).fill  = _fill(bg)
            ws.cell(row=row_i, column=3).font      = _font(size=9, italic=True, color="6B7280")
            ws.cell(row=row_i, column=3).border    = BORDER
            row_i += 1

        row_i += 1   # spacer row

    _set_col_width(ws, 1, 28)
    _set_col_width(ws, 2, 16)
    _set_col_width(ws, 3, 50)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 5: Monthly Returns
# ══════════════════════════════════════════════════════════════════════════════

def _make_monthly_returns(wb: Workbook, results: dict):
    ws = wb.create_sheet("Monthly Returns")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:N1")
    c = ws["A1"]
    c.value     = "Monthly Returns by Strategy  (%)"
    c.font      = _font(bold=True, color="FFFFFF", size=13)
    c.fill      = _fill("1B2A4A")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    row_i = 3
    for strat, res in results.items():
        r = res["daily_returns"]
        r.index = pd.DatetimeIndex(r.index)
        monthly = r.resample("ME").apply(lambda x: (1 + x).prod() - 1)
        monthly_df = monthly.to_frame("ret")
        monthly_df["Year"]  = monthly_df.index.year
        monthly_df["Month"] = monthly_df.index.month

        pivot = monthly_df.pivot(index="Year", columns="Month", values="ret")
        pivot.columns = [MONTHS[m - 1] for m in pivot.columns]
        # annual total
        pivot["Annual"] = (1 + monthly_df.groupby("Year")["ret"].apply(
            lambda x: (1 + x).prod() - 1)).values

        # Strategy header
        ws.merge_cells(f"A{row_i}:N{row_i}")
        sh = ws.cell(row=row_i, column=1, value=f"◆  {strat}")
        sh.font      = _font(bold=True, color="FFFFFF", size=11)
        sh.fill      = _fill("2563EB")
        sh.alignment = _align()
        row_i += 1

        # Column headers
        col_hdrs = ["Year"] + MONTHS + ["Annual"]
        _header_row(ws, row_i, col_hdrs, fill_color="DBEAFE", font_color="1B2A4A")
        row_i += 1

        data_start = row_i
        for yr, row in pivot.iterrows():
            ws.cell(row=row_i, column=1, value=str(yr)).font   = _font(bold=True)
            ws.cell(row=row_i, column=1).alignment = _align("center")
            ws.cell(row=row_i, column=1).border    = BORDER
            for col_i, mth in enumerate(col_hdrs[1:], start=2):
                val = row.get(mth, np.nan)
                c = ws.cell(row=row_i, column=col_i,
                            value=round(float(val) * 100, 2) if not np.isnan(val) else "")
                c.number_format = '0.00'
                c.border        = BORDER
                c.alignment     = _align("center")
                c.font          = _font(size=9)
                if isinstance(val, float) and not np.isnan(val):
                    if val > 0:
                        c.fill = _fill("DCFCE7")
                        c.font = _font(size=9, color="16A34A")
                    elif val < 0:
                        c.fill = _fill("FEE2E2")
                        c.font = _font(size=9, color="DC2626")
            row_i += 1

        row_i += 2   # spacer

    _set_col_width(ws, 1, 8)
    for i in range(2, 15):
        _set_col_width(ws, i, 9)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 6: Drawdown Analysis
# ══════════════════════════════════════════════════════════════════════════════

def _make_drawdown(wb: Workbook, results: dict):
    ws = wb.create_sheet("Drawdown Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "Drawdown Analysis — All Strategies"
    c.font      = _font(bold=True, color="FFFFFF", size=13)
    c.fill      = _fill("1B2A4A")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    headers = ["Date"] + [f"{s} DD (%)" for s in results]
    _header_row(ws, 3, headers)

    # Build drawdown series
    dd_series = {}
    for strat, res in results.items():
        eq = res["equity_curve"]
        peak = eq.cummax()
        dd_series[strat] = ((eq - peak) / peak * 100).round(2)

    df = pd.DataFrame(dd_series)
    df.index = pd.DatetimeIndex(df.index).normalize()

    for row_i, (date, row) in enumerate(df.iterrows(), start=4):
        ws.cell(row=row_i, column=1, value=date.date()).border = BORDER
        ws.cell(row=row_i, column=1).fill      = _fill("F9FAFB")
        ws.cell(row=row_i, column=1).font      = _font(size=9)
        ws.cell(row=row_i, column=1).alignment = _align("center")
        for col_i, val in enumerate(row, start=2):
            c = ws.cell(row=row_i, column=col_i, value=float(val))
            c.number_format = "0.00"
            c.border        = BORDER
            c.alignment     = _align("right")
            c.font          = _font(size=9, color="DC2626" if val < -5 else "111827")
            c.fill          = _fill("FEE2E2" if val < -10 else
                                    "FEF3C7" if val < -5 else
                                    "FFFFFF" if row_i % 2 == 0 else "F9FAFB")

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 14
    for i in range(2, len(headers) + 1):
        _set_col_width(ws, i, 20)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 7: Paper Trading — Session Overview
# ══════════════════════════════════════════════════════════════════════════════

def _make_paper_overview(wb: Workbook, paper_results: dict):
    ws = wb.create_sheet("PT — Session Overview")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "📡  Paper Trading — Live Session Overview"
    c.font      = _font(bold=True, color="FFFFFF", size=14)
    c.fill      = _fill("1B2A4A")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value     = "Simulated live execution with market-impact slippage, per-trade commissions & daily risk guardrails"
    sub.font      = _font(italic=True, color="6B7280", size=9)
    sub.fill      = _fill("F9FAFB")
    sub.alignment = _align("center")

    # ── Summary KPI cards ─────────────────────────────────────────────────────
    row_i = 4
    for strat_name, res in paper_results.items():
        summary = res["summary"]

        ws.merge_cells(f"A{row_i}:F{row_i}")
        sh = ws.cell(row=row_i, column=1, value=f"◆  {strat_name}")
        sh.font      = _font(bold=True, color="FFFFFF", size=11)
        sh.fill      = _fill("2563EB")
        sh.alignment = _align()
        row_i += 1

        kpi_pairs = list(summary.items())
        # 2 KPIs per row
        for i in range(0, len(kpi_pairs), 2):
            left_k,  left_v  = kpi_pairs[i]
            right_k, right_v = kpi_pairs[i + 1] if i + 1 < len(kpi_pairs) else ("", "")
            bg = "F9FAFB" if row_i % 2 == 0 else "FFFFFF"

            lk = ws.cell(row=row_i, column=1, value=left_k)
            lk.font = _font(bold=True, size=9); lk.fill = _fill("DBEAFE"); lk.border = BORDER; lk.alignment = _align("right")
            lv = ws.cell(row=row_i, column=2, value=left_v)
            lv.font = _font(size=9, color="2563EB", bold=True); lv.fill = _fill(bg); lv.border = BORDER; lv.alignment = _align("center")
            ws.merge_cells(f"B{row_i}:C{row_i}")

            if right_k:
                rk = ws.cell(row=row_i, column=4, value=right_k)
                rk.font = _font(bold=True, size=9); rk.fill = _fill("DBEAFE"); rk.border = BORDER; rk.alignment = _align("right")
                rv = ws.cell(row=row_i, column=5, value=right_v)
                rv.font = _font(size=9, color="2563EB", bold=True); rv.fill = _fill(bg); rv.border = BORDER; rv.alignment = _align("center")
                ws.merge_cells(f"E{row_i}:F{row_i}")
            row_i += 1

        row_i += 1  # spacer

    # ── Equity curve table ────────────────────────────────────────────────────
    row_i += 1
    _section_title(ws, row_i, 1, "Daily Portfolio Value — All Strategies")
    row_i += 1

    col_headers = ["Date", "Cash", "Unrealised P&L", "Portfolio Value", "Daily P&L", "Open Positions"]
    strat_colors = ["2563EB", "16A34A", "DC2626"]

    for s_idx, (strat_name, res) in enumerate(paper_results.items()):
        eq = res["equity_curve"]
        offset = s_idx * (len(col_headers) + 1)

        hdr_row = row_i
        for c_i, h in enumerate(col_headers, start=1 + offset):
            cell = ws.cell(row=hdr_row, column=c_i, value=h if c_i == 1 + offset else h)
            cell.font      = _font(bold=True, color="FFFFFF", size=9)
            cell.fill      = _fill("1B2A4A")
            cell.border    = BORDER
            cell.alignment = _align("center")
        # strategy label above headers
        ws.merge_cells(
            start_row=hdr_row - 1, end_row=hdr_row - 1,
            start_column=1 + offset, end_column=len(col_headers) + offset
        )
        lbl = ws.cell(row=hdr_row - 1, column=1 + offset, value=strat_name)
        lbl.font      = _font(bold=True, color="FFFFFF", size=10)
        lbl.fill      = _fill(strat_colors[s_idx % len(strat_colors)])
        lbl.alignment = _align("center")

        for r_i, (_, row) in enumerate(eq.iterrows(), start=hdr_row + 1):
            bg = "FFFFFF" if r_i % 2 == 0 else "F9FAFB"
            for c_i, col in enumerate(col_headers, start=1 + offset):
                val = row[col] if col in row.index else ""
                cell = ws.cell(row=r_i, column=c_i, value=val)
                cell.fill      = _fill(bg)
                cell.border    = BORDER
                cell.font      = _font(size=9)
                cell.alignment = _align("center")
                if col in ("Cash", "Portfolio Value", "Daily P&L", "Unrealised P&L"):
                    cell.number_format = '$#,##0.00'
                    if col == "Daily P&L" and isinstance(val, (int, float)):
                        cell.font = _font(size=9, color="16A34A" if val >= 0 else "DC2626")

    _set_col_width(ws, 1, 13)
    for i in range(2, 7):
        _set_col_width(ws, i, 16)
    ws.freeze_panes = f"A{row_i + 1}"


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 8: Paper Trading — Order Execution Log
# ══════════════════════════════════════════════════════════════════════════════

def _make_paper_executions(wb: Workbook, paper_results: dict):
    ws = wb.create_sheet("PT — Execution Log")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "Order Execution Log — All Paper Trading Sessions"
    c.font      = _font(bold=True, color="FFFFFF", size=13)
    c.fill      = _fill("1B2A4A")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    headers = ["Strategy", "Date", "Ticker", "Side", "Type", "Qty", "Fill Price ($)", "Commission ($)", "Order ID"]
    _header_row(ws, 3, headers)

    row_i = 4
    for strat_name, res in paper_results.items():
        exec_log = res["execution_log"]
        if exec_log.empty:
            continue
        for _, order in exec_log.iterrows():
            is_buy = order.get("Side", "") == "BUY"
            bg = "DBEAFE" if (row_i % 2 == 0 and is_buy) else \
                 "DCFCE7" if (row_i % 2 == 0 and not is_buy) else \
                 "EFF6FF" if is_buy else "F0FDF4"

            vals = [strat_name, order.get("Date",""), order.get("Ticker",""),
                    order.get("Side",""), order.get("Type",""), order.get("Qty",""),
                    order.get("Fill Price",0), order.get("Commission",0), order.get("Order ID","")]

            for c_i, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_i, column=c_i, value=val)
                cell.fill      = _fill(bg)
                cell.border    = BORDER
                cell.alignment = _align("center")
                cell.font      = _font(size=9, color="2563EB" if is_buy else "16A34A", bold=(c_i == 4))
                if c_i == 7:
                    cell.number_format = '$#,##0.0000'
                elif c_i == 8:
                    cell.number_format = '$#,##0.00'
            row_i += 1

    ws.freeze_panes = "A4"
    widths = [22, 12, 8, 7, 9, 7, 16, 16, 12]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 9: Paper Trading — Live Positions
# ══════════════════════════════════════════════════════════════════════════════

def _make_paper_positions(wb: Workbook, paper_results: dict):
    ws = wb.create_sheet("PT — Live Positions")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "Live Positions — End-of-Session Snapshot"
    c.font      = _font(bold=True, color="FFFFFF", size=13)
    c.fill      = _fill("1B2A4A")
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    row_i = 3
    for strat_name, res in paper_results.items():
        positions = res["positions"]

        ws.merge_cells(f"A{row_i}:F{row_i}")
        sh = ws.cell(row=row_i, column=1, value=f"◆  {strat_name}")
        sh.font      = _font(bold=True, color="FFFFFF", size=11)
        sh.fill      = _fill("2563EB")
        sh.alignment = _align()
        row_i += 1

        if positions.empty:
            ws.cell(row=row_i, column=1, value="No open positions at session end.").font = _font(italic=True, color="6B7280")
            row_i += 2
            continue

        headers = ["Ticker", "Open Qty", "Avg Cost ($)", "Realised P&L ($)", "Unrealised P&L ($)", "Net P&L ($)"]
        _header_row(ws, row_i, headers, fill_color="DBEAFE", font_color="1B2A4A")
        row_i += 1

        for _, pos in positions.iterrows():
            net = pos.get("Realised P&L ($)", 0) + pos.get("Unrealised P&L ($)", 0)
            bg = "DCFCE7" if net >= 0 else "FEE2E2"
            vals = [
                pos["Ticker"], pos["Open Qty"],
                pos.get("Avg Cost", 0),
                pos.get("Realised P&L ($)", 0),
                pos.get("Unrealised P&L ($)", 0),
                round(net, 2),
            ]
            for c_i, val in enumerate(vals, start=1):
                cell = ws.cell(row=row_i, column=c_i, value=val)
                cell.fill      = _fill(bg)
                cell.border    = BORDER
                cell.alignment = _align("center")
                cell.font      = _font(size=9, color="16A34A" if (isinstance(val, float) and val >= 0 and c_i > 2)
                                                           else "DC2626" if (isinstance(val, float) and val < 0 and c_i > 2)
                                                           else "111827")
                if c_i in (3, 4, 5, 6):
                    cell.number_format = '$#,##0.00'
            row_i += 1

        # Summary totals row
        total_realised   = positions["Realised P&L ($)"].sum()   if "Realised P&L ($)"   in positions.columns else 0
        total_unrealised = positions["Unrealised P&L ($)"].sum() if "Unrealised P&L ($)" in positions.columns else 0
        total_net = total_realised + total_unrealised
        totals = ["TOTAL", "", "", round(total_realised,2), round(total_unrealised,2), round(total_net,2)]
        for c_i, val in enumerate(totals, start=1):
            cell = ws.cell(row=row_i, column=c_i, value=val)
            cell.font      = _font(bold=True, color="FFFFFF", size=9)
            cell.fill      = _fill("16A34A" if total_net >= 0 else "DC2626")
            cell.border    = BORDER
            cell.alignment = _align("center")
            if c_i in (4, 5, 6):
                cell.number_format = '$#,##0.00'

        row_i += 2

    widths = [12, 11, 14, 18, 20, 14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)
